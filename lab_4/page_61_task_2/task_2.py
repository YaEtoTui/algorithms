import openpyxl

# Создание нового Excel файла и листа
wb = openpyxl.Workbook()
ws = wb.active

# Заголовки
ws['A1'] = 'ξ2'
ws['B1'] = 'φ(ξ2)'
ws['C1'] = 'ψ(ξ2)'
ws['D1'] = 'Z1*(ξ2)'
ws['E1'] = 'Z2*(ξ2)'
ws['F1'] = 'Z3*(ξ2)'

# Ввод значений и формул
for i in range(0, 151, 5):
    row = (i // 5) + 2
    ws[f'A{row}'] = i
    ws[f'B{row}'] = f'=IF(A{row}<50, 0.1*A{row}, IF(A{row}<=150, (6/12 + 0.04*A{row}), 16))'
    ws[f'C{row}'] = f'=IF(A{row}<=100, 0.05*A{row} + 5, 0.09*A{row} + 31)'
    ws[f'D{row}'] = f'=IF(A{row}<=100, 0.09*A{row} + 20.5, 0.21*A{row} + 32.5)'
    ws[f'E{row}'] = f'=IF(A{row}<=50, 0.19*A{row} + 25.5, IF(A{row}<=100, 0.09*A{row} + 31, 0.09*A{row} + 31))'
    ws[f'F{row}'] = f'=IF(A{row}<=30, 0.3*A{row} + 74, 0.09*A{row} + 40.5)'

# Сохранение файла
wb.save('output.xlsx')