import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Исходные данные
n = 4
xi_0 = 0
xi_1 = 0
psi = [5 + 2 * i for i in range(1, 5)]
d = [6, 5, 15, 20]
c = 10

# Функция для вычисления xi
def calculate_xi(psi, d):
    return [(psi[i] - d[i]) / 2 for i in range(n)]

# Функция для вычисления функции затрат
def calculate_f(xi, x, d):
    return [(xi[i] - x[i] + d[i]) for i in range(n)]

# Вычисление xi
xi = calculate_xi(psi, d)

# Вычисление суммарных затрат
def calculate_Z(xi, x, d):
    f_values = calculate_f(xi, x, d)
    return sum(f_values)

# Оптимизация затрат
def optimize_Z(xi, d):
    x_opt = [min(0, xi[i] - d[i]) for i in range(n)]
    return x_opt, calculate_Z(xi, x_opt, d)

# Вычисление оптимальных значений
x_opt, Z_opt = optimize_Z(xi, d)

# Создание DataFrame для хранения результатов
data = {
    'k': [i + 1 for i in range(n)],
    'psi(xi)': psi,
    'd': d,
    'xi': xi,
    'x_opt': x_opt,
    'f(xi, x_opt)': calculate_f(xi, x_opt, d)
}

df = pd.DataFrame(data)

# Сохранение DataFrame в Excel
file_path = 'result.xlsx'
df.to_excel(file_path, index=False)

# Добавление формул в Excel
wb = load_workbook(file_path)
ws = wb.active

# Добавление формул для вычисления xi
for i in range(n):
    ws[f'F{i+2}'] = f'=(B{i+2}-C{i+2})/2'

# Добавление формул для вычисления функции затрат
for i in range(n):
    ws[f'G{i+2}'] = f'=(E{i+2}-F{i+2}+C{i+2})'

# Добавление суммарных затрат
ws['H1'] = 'Суммарные затраты'
ws['H2'] = Z_opt

# Сохранение изменений
wb.save(file_path)

print(f'Результаты сохранены в файл {file_path}')
