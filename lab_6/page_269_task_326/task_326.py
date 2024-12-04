import cvxpy as cp
from openpyxl import Workbook

# Определяем переменные
x1 = cp.Variable()
x2 = cp.Variable()
x3 = cp.Variable()

# Определяем целевую функцию
objective = cp.Maximize(-x1**2 - x2**2 - 2*x3**2 + 2*x2 + 3*x3)

# Определяем ограничения
constraints = [
    x1 + x2 + x3 <= 18,
    x2 <= 12,
    x1 + 2*x3 <= 14,
    x1 >= 0,
    x2 >= 0,
    x3 >= 0
]

# Определяем проблему
problem = cp.Problem(objective, constraints)

# Решаем проблему
problem.solve()

# Получаем значения переменных
x1_value = x1.value
x2_value = x2.value
x3_value = x3.value
max_value = problem.value

# Преобразуем значения в обычные числа
x1_value = float(x1_value) if x1_value is not None else 0
x2_value = float(x2_value) if x2_value is not None else 0
x3_value = float(x3_value) if x3_value is not None else 0
max_value = float(max_value) if max_value is not None else 0

# Выводим результат в консоль
print(f"x1 = {x1_value}")
print(f"x2 = {x2_value}")
print(f"x3 = {x3_value}")
print(f"Максимальное значение функции = {max_value}")

# Создаем новую книгу Excel
wb = Workbook()
ws = wb.active

# Записываем результаты в таблицу
ws['A1'] = "x1"
ws['B1'] = x1_value
ws['A2'] = "x2"
ws['B2'] = x2_value
ws['A3'] = "x3"
ws['B3'] = x3_value
ws['A4'] = "Максимальное значение функции"
ws['B4'] = max_value

# Сохраняем файл
wb.save("result.xlsx")
